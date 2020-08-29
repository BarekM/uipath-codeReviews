import os


import openpyxl
from openpyxl.utils import get_column_letter

class ExcelHandler():
    
    def __init__(self, path_excel):
        self.path_excel = path_excel
        self.__init_file()
        
    def __del__(self):
        self.__save_wb()
    
    def __save_wb(self):
        self.wb.save(filename=self.path_excel)
    
    def __init_sheet(self, title):
        if title in self.wb.sheetnames:
            return self.wb[title]
        else:
            return self.wb.create_sheet(title)
    
    def save_data(self, sheet_name, data, flag_autosize_columns=True):
        ws = self.__init_sheet(sheet_name)
        for d in data:
            ws.append(d)
        if flag_autosize_columns:
            self.__autosize_columns(ws, data)
        self.__save_wb()
    
    def __init_file(self):
         if os.path.exists(self.path_excel):
             self.wb = openpyxl.load_workbook(filename=self.path_excel)
         else:
             self.wb = openpyxl.Workbook()
             self.__save_wb
    
    def __autosize_columns(self, ws, data):
        column_widths = []
        for row in data:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell)) > column_widths[i]:
                        column_widths[i] = len(str(cell))
                else:
                    column_widths += [len(cell)]
        
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width
             
             
if __name__ == '__main__':
    p = r'C:\Users\markb\Documents\projects\uipath-codeReview\tests\test11.xlsx'
    
    eh = ExcelHandler(p)